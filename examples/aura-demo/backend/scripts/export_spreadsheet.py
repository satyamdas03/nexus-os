"""Export the synthetic 34k ASSURE book to a multi-sheet Excel workbook.

Sheets produced:
  - Portfolios    : 34,000 rows, one portfolio per row
  - Holdings      : every holding row with metadata + day-0 market value
  - Mandates      : the ~8 deduplicated mandate specs, expanded
  - Price Master  : every ticker with day-0 synthetic price + GBM params
  - Breach Reg    : computed rule breaches for each non-green portfolio at day 0
  - Book Summary  : single-row aggregate from status_history day 0

No real client data. All prices are synthetic GBM outputs.
"""
import json
import os
import sqlite3
import sys
from pathlib import Path

# so this script can be run from backend/ or project root
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from generators import universe as U
from core import rules_engine

DB_PATH = ROOT / "data" / "portfolios.db"
OUT_PATH = Path(__file__).resolve().parent.parent.parent / "ASSURE_Synthetic_34k_Book.xlsx"


def _connect():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def _sheet_portfolios(ws, conn):
    headers = [
        "client_id", "client_name", "adviser", "fum", "cash", "mandate_id",
        "day0_status", "breach_count", "watch_count",
    ]
    ws.append(headers)
    status = {
        r["client_id"]: dict(r)
        for r in conn.execute(
            "SELECT client_id, status, breach_count, watch_count FROM status_history WHERE day=0"
        )
    }
    for row in conn.execute(
        "SELECT client_id, client_name, adviser, fum, mandate_id, cash FROM portfolios ORDER BY client_id"
    ):
        s = status.get(row["client_id"], {})
        ws.append([
            row["client_id"], row["client_name"], row["adviser"],
            row["fum"], row["cash"], row["mandate_id"],
            s.get("status"), s.get("breach_count", 0), s.get("watch_count", 0),
        ])


def _sheet_holdings(ws, conn):
    headers = [
        "client_id", "ticker", "name", "asset_class", "sector", "region",
        "liquidity_tier", "units", "day0_price", "market_value",
    ]
    ws.append(headers)
    prices = {r["ticker"]: r["price"] for r in conn.execute("SELECT ticker, price FROM prices WHERE day=0")}
    for row in conn.execute(
        "SELECT client_id, ticker, units FROM holdings WHERE ticker!='CASH' ORDER BY client_id, ticker"
    ):
        meta = U.UNIVERSE_BY_TICKER.get(row["ticker"], {})
        price = prices.get(row["ticker"], 0.0)
        ws.append([
            row["client_id"], row["ticker"], meta.get("name"),
            meta.get("asset_class"), meta.get("sector"), meta.get("region"),
            meta.get("liquidity_tier"), row["units"], price,
            round(row["units"] * price, 2) if price else 0.0,
        ])


def _sheet_mandates(ws, conn):
    # Expanded JSON specs into a flat table.
    headers = [
        "mandate_id", "name",
        "max_asset_class_weight:Equity", "max_asset_class_weight:Bonds",
        "max_asset_class_weight:Commodity", "max_asset_class_weight:Crypto",
        "max_sector_weight_json", "max_region_weight:US", "max_region_weight:ExUS",
        "max_region_weight:EM", "max_single_holding", "min_cash",
        "target_allocation_json", "drift_tolerance",
        "approved_universe_csv", "excluded_tickers_csv",
        "top_n", "top_n_limit", "min_liquid_pct",
    ]
    ws.append(headers)
    for row in conn.execute("SELECT mandate_id, spec FROM mandates ORDER BY mandate_id"):
        m = json.loads(row["spec"])
        ac = m.get("max_asset_class_weight", {})
        reg = m.get("max_region_weight", {})
        topn = m.get("max_top_n_concentration", {})
        ws.append([
            row["mandate_id"], m.get("name"),
            ac.get("Equity"), ac.get("Bonds"), ac.get("Commodity"), ac.get("Crypto"),
            json.dumps(m.get("max_sector_weight")),
            reg.get("US"), reg.get("ExUS"), reg.get("EM"),
            m.get("max_single_holding"), m.get("min_cash"),
            json.dumps(m.get("target_allocation")), m.get("drift_tolerance"),
            ",".join(m.get("approved_universe", [])),
            ",".join(m.get("excluded_tickers", [])),
            topn.get("n"), topn.get("limit"), m.get("min_liquid_pct"),
        ])


def _sheet_prices(ws, conn):
    headers = ["ticker", "name", "asset_class", "sector", "region", "liquidity_tier",
               "day0_price", "base_price", "mu", "sigma"]
    ws.append(headers)
    prices = {r["ticker"]: r["price"] for r in conn.execute("SELECT ticker, price FROM prices WHERE day=0")}
    for u in U.UNIVERSE:
        ws.append([
            u["ticker"], u["name"], u["asset_class"], u["sector"], u["region"],
            u["liquidity_tier"], prices.get(u["ticker"], u["base_price"]),
            u["base_price"], u["mu"], u["sigma"],
        ])


def _sheet_breach_register(ws, conn):
    headers = [
        "client_id", "rule", "limit", "actual", "offending_holdings",
        "severity", "mandate_id",
    ]
    ws.append(headers)
    prices = {r["ticker"]: r["price"] for r in conn.execute("SELECT ticker, price FROM prices WHERE day=0")}
    for p in conn.execute("SELECT * FROM portfolios"):
        mandate = json.loads(
            conn.execute("SELECT spec FROM mandates WHERE mandate_id=?", (p["mandate_id"],)).fetchone()["spec"]
        )
        holdings = []
        for h in conn.execute(
            "SELECT ticker, units FROM holdings WHERE client_id=? AND ticker!='CASH'", (p["client_id"],)
        ):
            meta = U.UNIVERSE_BY_TICKER.get(h["ticker"])
            if not meta:
                continue
            price = prices[h["ticker"]]
            holdings.append({
                "ticker": h["ticker"], "units": h["units"], "price": price,
                "market_value": round(h["units"] * price, 2),
                "asset_class": meta["asset_class"], "sector": meta["sector"],
                "region": meta["region"], "liquidity_tier": meta["liquidity_tier"],
            })
        port = {
            "client_id": p["client_id"], "client_name": p["client_name"],
            "adviser": p["adviser"], "fum": p["fum"], "cash": p["cash"],
            "holdings": holdings, "mandate": mandate,
        }
        rr = rules_engine.check(port, mandate)
        if rr["status"] == "green":
            continue
        for b in rr.get("breaches", []):
            ws.append([
                p["client_id"], b["rule"], b.get("limit"), b.get("actual"),
                ",".join(str(x) for x in b.get("offending_holdings", [])),
                "red", p["mandate_id"],
            ])
        for w in rr.get("watches", []):
            ws.append([
                p["client_id"], w["rule"], w.get("limit"), w.get("actual"),
                ",".join(str(x) for x in w.get("offending_holdings", [])),
                "orange", p["mandate_id"],
            ])


def _sheet_summary(ws, conn):
    headers = ["day", "total", "green", "orange", "red", "breach_count"]
    ws.append(headers)
    row = conn.execute(
        "SELECT day, total, green, orange, red, breach_count FROM book_summary WHERE id=1"
    ).fetchone()
    if row:
        ws.append([row["day"], row["total"], row["green"], row["orange"], row["red"], row["breach_count"]])


def main():
    from openpyxl import Workbook
    from openpyxl.styles import Font

    if not DB_PATH.exists():
        print(f"DB not found at {DB_PATH}")
        sys.exit(1)

    conn = _connect()
    wb = Workbook()
    # Remove default sheet and add ordered sheets.
    wb.remove(wb.active)

    sheets = [
        ("README", _readme),
        ("Portfolios", _sheet_portfolios),
        ("Holdings", _sheet_holdings),
        ("Mandates", _sheet_mandates),
        ("Price Master", _sheet_prices),
        ("Breach Register", _sheet_breach_register),
        ("Book Summary", _sheet_summary),
    ]

    for name, writer in sheets:
        ws = wb.create_sheet(title=name)
        writer(ws, conn)
        # Freeze header row and bold it.
        if ws.max_row > 0:
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.freeze_panes = "A2"
        print(f"  sheet '{name}' -> {ws.max_row - 1} data rows")

    conn.close()
    wb.save(OUT_PATH)
    print(f"\nSaved: {OUT_PATH}")


def _readme(ws, conn):
    notes = [
        ["ASSURE Synthetic 34,000-Portfolio Book"],
        ["All data is synthetic. No real clients, no real prices."],
        ["Prices are generated by a seeded GBM model at day 0."],
        ["Portfolios, mandates, holdings, and breach statuses are deterministic given seed=42."],
        [],
        ["Sheet descriptions:"],
        ["Portfolios", "One row per portfolio with adviser, FUM, cash, and day-0 compliance status."],
        ["Holdings", "One row per holding with sector, region, liquidity tier, and market value."],
        ["Mandates", "The ~8 deduplicated mandate templates that the 34k portfolios share."],
        ["Price Master", "Reference ticker list with synthetic day-0 prices and GBM parameters."],
        ["Breach Register", "Computed rule breaches/watches for every non-green portfolio."],
        ["Book Summary", "Aggregate day-0 counts from the application database."],
    ]
    for r in notes:
        ws.append(r)


if __name__ == "__main__":
    main()
